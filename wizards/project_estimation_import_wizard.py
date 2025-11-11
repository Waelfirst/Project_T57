# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import logging

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    _logger.warning('openpyxl library not found')
    openpyxl = None


class ProjectEstimationImportWizard(models.TransientModel):
    _name = 'project.estimation.import.wizard'
    _description = 'Import Updated Cost Estimation'

    estimation_id = fields.Many2one('project.cost.estimation', string='Estimation', required=True)
    project_id = fields.Many2one('project.definition', string='Project', required=True)

    excel_file = fields.Binary(string='Updated Excel File', required=True, help='Upload the edited Excel file')
    filename = fields.Char(string='Filename')

    test_mode = fields.Boolean(string='Test Mode (Preview Only)', default=False,
                               help='Enable to preview data without importing')
    preview_data = fields.Text(string='Preview Data', readonly=True)
    update_summary = fields.Text(string='Update Summary', readonly=True)

    def action_import_prices(self):
        """Import updated prices from Excel file"""
        self.ensure_one()

        if not self.excel_file:
            raise UserError(_('Please upload an Excel file!'))

        if not openpyxl:
            raise UserError(_('Excel library not installed. Please contact administrator.'))

        try:
            file_data = base64.b64decode(self.excel_file)
            workbook = load_workbook(io.BytesIO(file_data), data_only=True)

            if 'Cost Estimation' not in workbook.sheetnames:
                available_sheets = ', '.join(workbook.sheetnames)
                raise UserError(
                    _('Invalid Excel file! Sheet "Cost Estimation" not found.\nAvailable sheets: %s') % available_sheets)

            ws = workbook['Cost Estimation']

            # Find data start row (after headers)
            data_start_row = None
            header_row_idx = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
                if row and len(row) > 0:
                    cell_value = str(row[0]).strip() if row[0] else ''
                    _logger.info('Row %s, Cell A: "%s"', row_idx, cell_value)
                    if cell_value == 'Product Code':
                        header_row_idx = row_idx
                        data_start_row = row_idx + 1
                        _logger.info('Found data header at row %s', row_idx)
                        break

            if not data_start_row:
                raise UserError(_('Could not find data header row!\n\nPlease ensure:\n'
                                  '1. Sheet name is "Cost Estimation"\n'
                                  '2. First column header is "Product Code"\n'
                                  '3. Header row exists in first 30 rows'))

            # Preview mode or actual import
            if self.test_mode:
                return self._preview_import(ws, data_start_row, header_row_idx)
            else:
                return self._execute_import(ws, data_start_row)

        except Exception as e:
            _logger.exception('Error processing Excel file')
            raise UserError(_('Error processing Excel file:\n\n%s\n\nCheck server logs for details.') % str(e))

    def _preview_import(self, ws, data_start_row, header_row_idx):
        """Preview what will be imported without making changes - ENHANCED"""
        preview_lines = []
        preview_lines.append('=' * 80)
        preview_lines.append('📋 IMPORT PREVIEW - NO CHANGES MADE')
        preview_lines.append('=' * 80)
        preview_lines.append('')

        # Show project info
        preview_lines.append(f'Project: {self.project_id.name}')
        preview_lines.append(f'Products in project: {len(self.project_id.product_line_ids)}')
        preview_lines.append('')

        # Show products in project
        preview_lines.append('Products available in project:')
        for pl in self.project_id.product_line_ids:
            code = pl.product_id.default_code or '(no code)'
            name = pl.product_id.name
            preview_lines.append(f'  • Code: {code} | Name: {name}')
        preview_lines.append('')

        # Show header row
        preview_lines.append(f'Header row found at: Row {header_row_idx}')
        preview_lines.append(f'Data starts at: Row {data_start_row}')
        preview_lines.append('')

        # NEW: Show actual header structure
        preview_lines.append('📋 EXCEL HEADER STRUCTURE:')
        preview_lines.append('-' * 80)
        header_row = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
        for col_idx in range(min(20, len(header_row))):
            col_letter = chr(65 + col_idx)
            header_val = header_row[col_idx] if col_idx < len(header_row) else 'N/A'
            preview_lines.append(f'  {col_letter} (idx {col_idx}): {header_val}')
        preview_lines.append('')

        # Highlight target columns
        preview_lines.append('🎯 TARGET COLUMNS FOR IMPORT:')
        if len(header_row) > 15:
            preview_lines.append(f'  Column P (idx 15): {header_row[15]}')
        else:
            preview_lines.append(f'  ❌ Column P NOT FOUND (file has only {len(header_row)} columns)')

        if len(header_row) > 16:
            preview_lines.append(f'  Column Q (idx 16): {header_row[16]}')
        else:
            preview_lines.append(f'  ❌ Column Q NOT FOUND (file has only {len(header_row)} columns)')
        preview_lines.append('')

        # Show first few data rows with MORE details
        preview_lines.append('📊 FIRST 5 DATA ROWS:')
        preview_lines.append('-' * 80)

        matched_count = 0
        unmatched_count = 0

        for row_idx, row in enumerate(
                ws.iter_rows(min_row=data_start_row, max_row=data_start_row + 4, values_only=True),
                start=data_start_row):
            if not row or not row[0]:
                continue

            product_code = str(row[0]).strip() if row[0] else ''
            product_name = str(row[1]).strip() if row[1] else ''

            # Try to find product
            product_line = self._find_product_line(product_code, product_name)

            # Read prices from Column P (index 15) and Column Q (index 16)
            cost_price = None
            sale_price = None
            cost_price_raw = 'N/A'
            sale_price_raw = 'N/A'

            if len(row) > 15:
                cost_price_raw = row[15]
                if row[15] is not None:
                    try:
                        cost_price = float(row[15])
                    except:
                        pass

            if len(row) > 16:
                sale_price_raw = row[16]
                if row[16] is not None:
                    try:
                        sale_price = float(row[16])
                    except:
                        pass

            preview_lines.append(f'\n📍 Row {row_idx}:')
            preview_lines.append(f'   Product Code (A): "{product_code}"')
            preview_lines.append(f'   Product Name (B): "{product_name}"')
            preview_lines.append(f'   Column P raw value: {cost_price_raw}')
            preview_lines.append(f'   Column P parsed: {cost_price if cost_price is not None else "FAILED"}')
            preview_lines.append(f'   Column Q raw value: {sale_price_raw}')
            preview_lines.append(f'   Column Q parsed: {sale_price if sale_price is not None else "FAILED"}')

            if product_line:
                matched_count += 1
                preview_lines.append(f'   ✅ MATCHED: {product_line.product_id.display_name}')
                preview_lines.append(
                    f'   Current Cost: {product_line.cost_price:.2f} → New: {cost_price if cost_price else "N/A"}')
                preview_lines.append(
                    f'   Current Sale: {product_line.sale_price:.2f} → New: {sale_price if sale_price else "N/A"}')
            else:
                unmatched_count += 1
                preview_lines.append(f'   ❌ NOT MATCHED - Product not found in project')

        preview_lines.append('')
        preview_lines.append('=' * 80)
        preview_lines.append(f'Preview Summary:')
        preview_lines.append(f'  Matched products: {matched_count}')
        preview_lines.append(f'  Unmatched products: {unmatched_count}')
        preview_lines.append('=' * 80)
        preview_lines.append('')
        preview_lines.append('💡 TO PROCEED WITH IMPORT:')
        preview_lines.append('   1. Uncheck "Test Mode (Preview Only)"')
        preview_lines.append('   2. Click "Import Prices" again')
        preview_lines.append('')
        preview_lines.append('⚠️  BEFORE IMPORTING:')
        preview_lines.append('   • Check that Column P and Q are detected correctly')
        preview_lines.append('   • Verify parsed values are correct (not N/A or FAILED)')
        preview_lines.append('   • Fix unmatched products (codes/names must match exactly)')
        preview_lines.append('   • If columns are wrong, regenerate the Excel file')

        self.preview_data = '\n'.join(preview_lines)

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def _find_product_line(self, product_code, product_name):
        """Find product line using multiple strategies"""
        if not product_code and not product_name:
            return False

        # Strategy 1: Exact match by code
        if product_code:
            product_line = self.project_id.product_line_ids.filtered(
                lambda l: l.product_id.default_code and l.product_id.default_code.strip() == product_code
            )
            if product_line:
                return product_line[0] if len(product_line) > 1 else product_line

        # Strategy 2: Exact match by name
        if product_name:
            product_line = self.project_id.product_line_ids.filtered(
                lambda l: l.product_id.name and l.product_id.name.strip() == product_name
            )
            if product_line:
                return product_line[0] if len(product_line) > 1 else product_line

        # Strategy 3: Case-insensitive match by name
        if product_name:
            product_line = self.project_id.product_line_ids.filtered(
                lambda l: l.product_id.name and l.product_id.name.strip().lower() == product_name.lower()
            )
            if product_line:
                return product_line[0] if len(product_line) > 1 else product_line

        return False

    def _execute_import(self, ws, data_start_row):
        """Execute the actual import - FIXED with correct column indices"""
        updates = []
        updated_count = 0
        errors = []
        skipped_count = 0

        _logger.info('=' * 80)
        _logger.info('STARTING IMPORT - ROW %s', data_start_row)
        _logger.info('=' * 80)

        # Debug: Show products in project
        _logger.info('Products in project:')
        for pl in self.project_id.product_line_ids:
            _logger.info('  - Code: %s | Name: %s | Cost: %.2f | Sale: %.2f',
                         pl.product_id.default_code or '(no code)',
                         pl.product_id.name,
                         pl.cost_price,
                         pl.sale_price)

        for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            # Stop at TOTAL row or empty row
            if not row or not row[0]:
                _logger.info('Row %s: Empty row, stopping', row_idx)
                break

            cell_a_value = str(row[0]).strip().upper() if row[0] else ''
            if cell_a_value == 'TOTAL':
                _logger.info('Row %s: TOTAL row found, stopping', row_idx)
                break

            product_code = str(row[0]).strip() if row[0] else ''
            product_name = str(row[1]).strip() if row[1] else ''

            _logger.info('-' * 80)
            _logger.info('Row %s: Processing', row_idx)
            _logger.info('  Product Code (Column A): "%s"', product_code)
            _logger.info('  Product Name (Column B): "%s"', product_name)

            # Read prices from Excel
            # Column P (index 15) = Cost Price (Calculated)
            # Column Q (index 16) = Sale Price (Editable)
            cost_price = None
            sale_price = None
            cost_col_idx = 15  # Column P
            sale_col_idx = 16  # Column Q

            # Read Cost Price from column P (index 15)
            if len(row) > 15 and row[15] is not None:
                try:
                    cost_price = float(row[15])
                    _logger.info('  Cost Price (Column P, index 15): %.2f', cost_price)
                except (ValueError, TypeError) as e:
                    _logger.warning('  Could not read cost price from column P: %s', e)

            # Read Sale Price from column Q (index 16)
            if len(row) > 16 and row[16] is not None:
                try:
                    sale_price = float(row[16])
                    _logger.info('  Sale Price (Column Q, index 16): %.2f', sale_price)
                except (ValueError, TypeError) as e:
                    _logger.warning('  Could not read sale price from column Q: %s', e)

            if not product_code and not product_name:
                _logger.info('  → SKIPPED: No product code or name')
                skipped_count += 1
                continue

            # Find product line
            _logger.info('  Searching for product...')
            product_line = self._find_product_line(product_code, product_name)

            if not product_line:
                error_msg = f'Row {row_idx}: Product not found - Code: "{product_code}", Name: "{product_name}"'
                errors.append(error_msg)
                _logger.warning('  → NOT FOUND: %s', error_msg)
                continue

            _logger.info('  → FOUND: %s (ID: %s)', product_line.product_id.name, product_line.product_id.id)

            # Update prices only if values were found in Excel
            old_cost = product_line.cost_price
            old_sale = product_line.sale_price

            _logger.info('  Old prices: Cost=%.2f, Sale=%.2f', old_cost, old_sale)

            update_vals = {}
            if cost_price is not None:
                update_vals['cost_price'] = cost_price
            if sale_price is not None:
                update_vals['sale_price'] = sale_price

            if update_vals:
                try:
                    product_line.write(update_vals)
                    _logger.info('  → UPDATED: %s', update_vals)

                    updates.append({
                        'product': product_name or product_line.product_id.name,
                        'old_cost': old_cost,
                        'new_cost': cost_price if cost_price is not None else old_cost,
                        'old_sale': old_sale,
                        'new_sale': sale_price if sale_price is not None else old_sale,
                    })
                    updated_count += 1
                except Exception as e:
                    error_msg = f'Row {row_idx}: Failed to update product "{product_name}": {str(e)}'
                    errors.append(error_msg)
                    _logger.error('  → ERROR: %s', error_msg)
            else:
                _logger.warning('  → NO UPDATE: No prices found in Excel')

        # Update estimation record
        self.estimation_id.write({
            'excel_file': self.excel_file,
            'excel_filename': self.filename,
            'last_update_date': fields.Datetime.now(),
            'state': 'completed',
        })

        _logger.info('=' * 80)
        _logger.info('IMPORT COMPLETED: %s products updated, %s errors, %s skipped',
                     updated_count, len(errors), skipped_count)
        _logger.info('=' * 80)

        # Create summary message
        summary = _('✅ Successfully updated %s products!\n') % updated_count

        if skipped_count > 0:
            summary += _('⏭️ Skipped %s empty rows\n') % skipped_count

        summary += '\n'

        if updates:
            summary += _('📊 Updated Products:\n')
            for update in updates[:10]:
                summary += _('• %s\n') % update['product']
                if update['new_cost'] != update['old_cost']:
                    summary += _('  Cost: %.2f → %.2f\n') % (update['old_cost'], update['new_cost'])
                if update['new_sale'] != update['old_sale']:
                    summary += _('  Sale: %.2f → %.2f\n') % (update['old_sale'], update['new_sale'])

            if len(updates) > 10:
                summary += _('\n... and %s more products\n') % (len(updates) - 10)

        if errors:
            summary += _('\n\n⚠️ Errors (%s):\n') % len(errors)
            for error in errors[:5]:
                summary += f'• {error}\n'
            if len(errors) > 5:
                summary += _('... and %s more errors\n') % (len(errors) - 5)

        # If nothing was updated, show helpful debug info
        if updated_count == 0:
            summary += _('\n\n🔍 Debug Information:\n')
            summary += _('Expected columns:\n')
            summary += _('  • Column P (index 15): Cost Price\n')
            summary += _('  • Column Q (index 16): Sale Price\n')
            summary += _('\nPlease check:\n')
            summary += _('  1. Excel file has correct column structure\n')
            summary += _('  2. Cost Price is in column P\n')
            summary += _('  3. Sale Price is in column Q\n')
            summary += _('  4. Product codes/names match exactly\n')
            summary += _('\nCheck server logs for detailed information.\n')

        # Post message to project
        self.project_id.message_post(
            body=_('📊 Cost estimation import attempted.<br/>'
                   'Updated: %s products<br/>'
                   'Errors: %s<br/>'
                   'Skipped: %s<br/>'
                   'Estimation: %s') % (updated_count, len(errors), skipped_count, self.estimation_id.name),
            subject=_('Cost Estimation Import')
        )

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import Complete') if updated_count > 0 else _('Import Failed'),
                'message': summary,
                'type': 'success' if updated_count > 0 and not errors else 'warning' if updated_count > 0 else 'danger',
                'sticky': True,
            }
        }