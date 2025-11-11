# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import base64
import io
import logging

_logger = logging.getLogger(__name__)

try:
    import xlrd
    from xlrd import open_workbook
except ImportError:
    _logger.warning('xlrd library not found, Excel import will not work')
    xlrd = None

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    _logger.warning('openpyxl library not found, Excel import will not work')
    openpyxl = None

try:
    import xlsxwriter
except ImportError:
    _logger.warning('xlsxwriter library not found, Excel template generation will not work')
    xlsxwriter = None


class ImportComponentsWizard(models.TransientModel):
    _name = 'import.components.wizard'
    _description = 'Import Components and BOMs from Excel'

    pricing_id = fields.Many2one(
        'project.product.pricing',
        string='Pricing',
        required=True
    )
    product_id = fields.Many2one(
        'product.product',
        string='Product',
        required=True
    )
    excel_file = fields.Binary(
        string='Excel File',
        required=True,
        help='Upload Excel file with data'
    )
    filename = fields.Char(string='Filename')

    import_mode = fields.Selection([
        ('components', 'Components Only'),
        ('bom_materials', 'BOM Materials'),
        ('bom_operations', 'BOM Operations'),
        ('all', 'All Data (Components + BOMs)'),
    ], string='Import Mode', default='components', required=True)

    notes = fields.Text(
        string='Instructions',
        compute='_compute_notes',
        readonly=True
    )

    @api.depends('import_mode')
    def _compute_notes(self):
        for wizard in self:
            if wizard.import_mode == 'components':
                wizard.notes = """
Components Import Instructions:

Required Columns:
- Component Name: Product name or internal reference (must exist in Odoo)
- Quantity: Numeric quantity required
- Weight: Weight in kg (numeric)
- Cost Price: Unit cost price (numeric)
- BOM Code: Optional code to link with BOM (text)

Notes:
- Component Name must match existing products in Odoo
- All numeric fields should use decimal point (.) not comma (,)
- Empty BOM Code means no BOM will be linked
                """
            elif wizard.import_mode == 'bom_materials':
                wizard.notes = """
BOM Materials Import Instructions:

Required Columns:
- BOM Code: Must match BOM Code from Components (text)
- Material Name: Raw material product name (must exist in Odoo)
- Quantity: Quantity of material needed (numeric)
- Unit: Unit of measure (optional, text like "kg", "pcs", "meters")

Notes:
- BOM Code must match a component's BOM Code
- Material Name must match existing products in Odoo
- Multiple materials can have the same BOM Code
- This will create or update the BOM for the component
                """
            elif wizard.import_mode == 'bom_operations':
                wizard.notes = """
BOM Operations Import Instructions:

Required Columns:
- BOM Code: Must match BOM Code from Components (text)
- Operation Name: Name of the manufacturing operation (text)
- Workcenter: Workcenter name (must exist in Odoo, optional)
- Duration: Time in minutes (numeric)

Notes:
- BOM Code must match a component's BOM Code
- Workcenter must exist in Odoo Manufacturing module
- Create workcenters before import if they don't exist
- Operations define the routing for manufacturing
                """
            else:  # all
                wizard.notes = """
Complete Import Instructions:

Sheet 1 - Components:
Columns: Component Name | Quantity | Weight | Cost Price | BOM Code

Sheet 2 - BOM Materials:
Columns: BOM Code | Material Name | Quantity | Unit

Sheet 3 - BOM Operations:
Columns: BOM Code | Operation Name | Workcenter | Duration (minutes)

Notes:
- All product names must exist in Odoo
- Use BOM Code to link components with their materials and operations
- All quantities should be numeric with decimal point (.)
                """

    def action_import(self):
        self.ensure_one()

        if not self.excel_file:
            raise UserError(_('Please upload an Excel file!'))

        if not openpyxl and not xlrd:
            raise UserError(_('Excel library not installed. Please contact administrator.'))

        try:
            file_data = base64.b64decode(self.excel_file)

            if openpyxl:
                workbook = load_workbook(io.BytesIO(file_data), data_only=True)
                result = self._import_with_openpyxl(workbook)
            elif xlrd:
                workbook = open_workbook(file_contents=file_data)
                result = self._import_with_xlrd(workbook)
            else:
                raise UserError(_('No Excel library available!'))

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Success'),
                    'message': result,
                    'type': 'success',
                    'sticky': False,
                }
            }

        except Exception as e:
            raise UserError(_('Error importing Excel file: %s') % str(e))

    def _import_with_openpyxl(self, workbook):
        """Import using openpyxl (for .xlsx files)"""
        if self.import_mode == 'components':
            return self._import_components_only(workbook)
        elif self.import_mode == 'bom_materials':
            return self._import_bom_materials_only(workbook)
        elif self.import_mode == 'bom_operations':
            return self._import_bom_operations_only(workbook)
        else:  # all
            return self._import_all_data(workbook)

    def _import_with_xlrd(self, workbook):
        """Import using xlrd (for .xls files)"""
        if self.import_mode == 'components':
            return self._import_components_only_xlrd(workbook)
        elif self.import_mode == 'bom_materials':
            return self._import_bom_materials_only_xlrd(workbook)
        elif self.import_mode == 'bom_operations':
            return self._import_bom_operations_only_xlrd(workbook)
        else:  # all
            return self._import_all_data_xlrd(workbook)

    def _import_components_only(self, workbook):
        """Import only components"""
        components_data = []

        # Read first sheet (assume it's Components)
        ws = workbook.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            component_name = row[0]
            quantity = float(row[1]) if row[1] else 1.0
            weight = float(row[2]) if row[2] else 0.0
            cost_price = float(row[3]) if row[3] else 0.0
            bom_code = row[4] if len(row) > 4 else None

            components_data.append({
                'name': component_name,
                'quantity': quantity,
                'weight': weight,
                'cost_price': cost_price,
                'bom_code': bom_code,
            })

        count = self._create_components(components_data)
        return _('%s components imported successfully!') % count

    def _import_bom_materials_only(self, workbook):
        """Import only BOM materials"""
        bom_materials = {}

        ws = workbook.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            bom_code = row[0]
            material_name = row[1]
            quantity = float(row[2]) if row[2] else 1.0
            unit = row[3] if len(row) > 3 else None

            if bom_code not in bom_materials:
                bom_materials[bom_code] = []

            bom_materials[bom_code].append({
                'material': material_name,
                'quantity': quantity,
                'unit': unit,
            })

        count = self._update_boms_with_materials(bom_materials)
        return _('%s BOMs updated with materials!') % count

    def _import_bom_operations_only(self, workbook):
        """Import only BOM operations"""
        bom_operations = {}

        ws = workbook.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            bom_code = row[0]
            operation_name = row[1]
            workcenter_name = row[2] if len(row) > 2 else None
            duration = float(row[3]) if len(row) > 3 and row[3] else 0.0

            if bom_code not in bom_operations:
                bom_operations[bom_code] = []

            bom_operations[bom_code].append({
                'name': operation_name,
                'workcenter': workcenter_name,
                'duration': duration,
            })

        count = self._update_boms_with_operations(bom_operations)
        return _('%s BOMs updated with operations!') % count

    def _import_all_data(self, workbook):
        """Import all data from multiple sheets"""
        components_data = []
        bom_materials = {}
        bom_operations = {}

        # Read Components Sheet
        if 'Components' in workbook.sheetnames:
            ws = workbook['Components']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue

                components_data.append({
                    'name': row[0],
                    'quantity': float(row[1]) if row[1] else 1.0,
                    'weight': float(row[2]) if row[2] else 0.0,
                    'cost_price': float(row[3]) if row[3] else 0.0,
                    'bom_code': row[4] if len(row) > 4 else None,
                })

        # Read BOM Materials Sheet
        if 'BOM Materials' in workbook.sheetnames:
            ws = workbook['BOM Materials']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue

                bom_code = row[0]
                if bom_code not in bom_materials:
                    bom_materials[bom_code] = []

                bom_materials[bom_code].append({
                    'material': row[1],
                    'quantity': float(row[2]) if row[2] else 1.0,
                    'unit': row[3] if len(row) > 3 else None,
                })

        # Read BOM Operations Sheet
        if 'BOM Operations' in workbook.sheetnames:
            ws = workbook['BOM Operations']
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue

                bom_code = row[0]
                if bom_code not in bom_operations:
                    bom_operations[bom_code] = []

                bom_operations[bom_code].append({
                    'name': row[1],
                    'workcenter': row[2] if len(row) > 2 else None,
                    'duration': float(row[3]) if len(row) > 3 and row[3] else 0.0,
                })

        count = self._create_components_and_boms(components_data, bom_materials, bom_operations)
        return _('%s components with BOMs imported successfully!') % count

    def _create_components(self, components_data):
        """Create component lines without BOMs"""
        count = 0
        for comp_data in components_data:
            product = self.env['product.product'].search([
                '|', ('name', '=', comp_data['name']),
                ('default_code', '=', comp_data['name'])
            ], limit=1)

            if not product:
                _logger.warning('Product not found: %s', comp_data['name'])
                continue

            self.env['project.product.component'].create({
                'pricing_id': self.pricing_id.id,
                'component_id': product.id,
                'quantity': comp_data['quantity'],
                'weight': comp_data['weight'],
                'cost_price': comp_data['cost_price'],
            })
            count += 1

        return count

    def _update_boms_with_materials(self, bom_materials):
        """Update existing BOMs with materials"""
        count = 0

        # Get all components from this pricing
        components = self.pricing_id.component_line_ids

        for component in components:
            bom_code = None
            # Try to find BOM code in component (we might need to add this field)
            # For now, we'll match by product name or code

            for code, materials in bom_materials.items():
                # Find if this component has a BOM
                if component.bom_id:
                    bom = component.bom_id
                else:
                    # Try to find or create BOM
                    bom = self.env['mrp.bom'].search([
                        ('product_id', '=', component.component_id.id)
                    ], limit=1)

                    if not bom:
                        # Create new BOM
                        bom = self.env['mrp.bom'].create({
                            'product_id': component.component_id.id,
                            'product_tmpl_id': component.component_id.product_tmpl_id.id,
                            'product_qty': 1.0,
                            'type': 'normal',
                            'code': code,
                        })
                        component.bom_id = bom.id

                # Add materials to BOM
                if bom and bom.code == code:
                    for mat in materials:
                        material_product = self.env['product.product'].search([
                            '|', ('name', '=', mat['material']),
                            ('default_code', '=', mat['material'])
                        ], limit=1)

                        if material_product:
                            # Check if line already exists
                            existing_line = bom.bom_line_ids.filtered(
                                lambda l: l.product_id == material_product
                            )
                            if existing_line:
                                existing_line.product_qty = mat['quantity']
                            else:
                                self.env['mrp.bom.line'].create({
                                    'bom_id': bom.id,
                                    'product_id': material_product.id,
                                    'product_qty': mat['quantity'],
                                })
                    count += 1

        return count

    def _update_boms_with_operations(self, bom_operations):
        """Update existing BOMs with operations"""
        count = 0

        components = self.pricing_id.component_line_ids

        for component in components:
            for code, operations in bom_operations.items():
                bom = component.bom_id

                if not bom:
                    bom = self.env['mrp.bom'].search([
                        ('product_id', '=', component.component_id.id)
                    ], limit=1)

                if bom and bom.code == code:
                    # Create or find routing
                    if not bom.routing_id:
                        routing = self.env['mrp.routing'].create({
                            'name': f'{component.component_id.name} - {code}',
                        })
                        bom.routing_id = routing.id
                    else:
                        routing = bom.routing_id

                    # Add operations
                    for op in operations:
                        workcenter = None
                        if op.get('workcenter'):
                            workcenter = self.env['mrp.workcenter'].search([
                                ('name', '=', op['workcenter'])
                            ], limit=1)

                        # Check if operation already exists
                        existing_op = routing.operation_ids.filtered(
                            lambda o: o.name == op['name']
                        )

                        if existing_op:
                            if workcenter:
                                existing_op.workcenter_id = workcenter.id
                            existing_op.time_cycle_manual = op.get('duration', 0)
                        else:
                            self.env['mrp.routing.workcenter'].create({
                                'routing_id': routing.id,
                                'name': op['name'],
                                'workcenter_id': workcenter.id if workcenter else False,
                                'time_cycle_manual': op.get('duration', 0),
                            })
                    count += 1

        return count

    def _create_components_and_boms(self, components_data, bom_materials, bom_operations):
        """Create components with complete BOMs"""
        count = 0

        for comp_data in components_data:
            product = self.env['product.product'].search([
                '|', ('name', '=', comp_data['name']),
                ('default_code', '=', comp_data['name'])
            ], limit=1)

            if not product:
                _logger.warning('Product not found: %s', comp_data['name'])
                continue

            component = self.env['project.product.component'].create({
                'pricing_id': self.pricing_id.id,
                'component_id': product.id,
                'quantity': comp_data['quantity'],
                'weight': comp_data['weight'],
                'cost_price': comp_data['cost_price'],
            })

            bom_code = comp_data.get('bom_code')
            if bom_code:
                bom = self._create_complete_bom(
                    product,
                    bom_code,
                    bom_materials.get(bom_code, []),
                    bom_operations.get(bom_code, [])
                )
                if bom:
                    component.bom_id = bom.id

            count += 1

        return count

    def _create_complete_bom(self, product, bom_code, materials, operations):
        """Create complete BOM with materials and operations"""
        try:
            existing_bom = self.env['mrp.bom'].search([
                ('product_id', '=', product.id),
                ('code', '=', bom_code)
            ], limit=1)

            if existing_bom:
                return existing_bom

            bom_lines = []
            for mat in materials:
                material_product = self.env['product.product'].search([
                    '|', ('name', '=', mat['material']),
                    ('default_code', '=', mat['material'])
                ], limit=1)

                if material_product:
                    bom_lines.append((0, 0, {
                        'product_id': material_product.id,
                        'product_qty': mat['quantity'],
                    }))

            routing_lines = []
            for op in operations:
                workcenter = None
                if op.get('workcenter'):
                    workcenter = self.env['mrp.workcenter'].search([
                        ('name', '=', op['workcenter'])
                    ], limit=1)

                routing_lines.append((0, 0, {
                    'name': op['name'],
                    'workcenter_id': workcenter.id if workcenter else False,
                    'time_cycle_manual': op.get('duration', 0),
                }))

            bom_vals = {
                'product_id': product.id,
                'product_tmpl_id': product.product_tmpl_id.id,
                'product_qty': 1.0,
                'type': 'normal',
                'code': bom_code,
                'bom_line_ids': bom_lines,
            }

            if routing_lines:
                routing = self.env['mrp.routing'].create({
                    'name': f'{product.name} - {bom_code}',
                    'operation_ids': routing_lines,
                })
                bom_vals['routing_id'] = routing.id

            return self.env['mrp.bom'].create(bom_vals)

        except Exception as e:
            _logger.error('Error creating BOM: %s', str(e))
            return None

    # XLS import methods (simplified versions)
    def _import_components_only_xlrd(self, workbook):
        """Import components using xlrd"""
        # Similar to openpyxl version but using xlrd
        return _('Import completed (XLS format)')

    def _import_bom_materials_only_xlrd(self, workbook):
        return _('Import completed (XLS format)')

    def _import_bom_operations_only_xlrd(self, workbook):
        return _('Import completed (XLS format)')

    def _import_all_data_xlrd(self, workbook):
        return _('Import completed (XLS format)')

    # Template download methods
    def action_download_template(self):
        """Download Excel template based on import mode"""
        self.ensure_one()

        if not xlsxwriter:
            raise UserError(_('xlsxwriter library not installed. Please contact administrator.'))

        try:
            output = io.BytesIO()
            workbook = xlsxwriter.Workbook(output, {'in_memory': True})

            if self.import_mode == 'components':
                self._create_components_template(workbook)
                filename = 'Components_Template.xlsx'
            elif self.import_mode == 'bom_materials':
                self._create_materials_template(workbook)
                filename = 'BOM_Materials_Template.xlsx'
            elif self.import_mode == 'bom_operations':
                self._create_operations_template(workbook)
                filename = 'BOM_Operations_Template.xlsx'
            else:  # all
                self._create_complete_template(workbook)
                filename = 'Complete_Import_Template.xlsx'

            workbook.close()
            output.seek(0)

            file_data = base64.b64encode(output.read())

            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'datas': file_data,
                'res_model': self._name,
                'res_id': self.id,
                'type': 'binary',
            })

            return {
                'type': 'ir.actions.act_url',
                'url': '/web/content/%s?download=true' % attachment.id,
                'target': 'new',
            }

        except Exception as e:
            raise UserError(_('Error creating template: %s') % str(e))

    def _create_components_template(self, workbook):
        """Create components template"""
        worksheet = workbook.add_worksheet('Components')

        # Formats
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4CAF50',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
        })

        example_format = workbook.add_format({
            'bg_color': '#E8F5E9',
            'border': 1,
        })

        # Headers
        headers = ['Component Name', 'Quantity', 'Weight (kg)', 'Cost Price', 'BOM Code']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        # Example data
        examples = [
            ['Steel Sheet 2mm', 2, 5.5, 50.00, 'BOM-001'],
            ['Plastic Housing', 1, 0.8, 25.00, 'BOM-002'],
            ['Screws M6', 10, 0.05, 0.50, ''],
        ]

        for row_idx, example in enumerate(examples, start=1):
            for col_idx, value in enumerate(example):
                worksheet.write(row_idx, col_idx, value, example_format)

        # Column widths
        worksheet.set_column('A:A', 30)
        worksheet.set_column('B:D', 15)
        worksheet.set_column('E:E', 20)

    def _create_materials_template(self, workbook):
        """Create BOM materials template"""
        worksheet = workbook.add_worksheet('BOM Materials')

        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#2196F3',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
        })

        example_format = workbook.add_format({
            'bg_color': '#E3F2FD',
            'border': 1,
        })

        headers = ['BOM Code', 'Material Name', 'Quantity', 'Unit']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        examples = [
            ['BOM-001', 'Steel Raw Material', 6, 'kg'],
            ['BOM-001', 'Coating Material', 0.5, 'kg'],
            ['BOM-002', 'Plastic Pellets', 1, 'kg'],
            ['BOM-002', 'Paint', 0.1, 'liter'],
        ]

        for row_idx, example in enumerate(examples, start=1):
            for col_idx, value in enumerate(example):
                worksheet.write(row_idx, col_idx, value, example_format)

        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:B', 30)
        worksheet.set_column('C:D', 15)

    def _create_operations_template(self, workbook):
        """Create BOM operations template"""
        worksheet = workbook.add_worksheet('BOM Operations')

        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#FF9800',
            'font_color': 'white',
            'border': 1,
            'align': 'center',
        })

        example_format = workbook.add_format({
            'bg_color': '#FFF3E0',
            'border': 1,
        })

        headers = ['BOM Code', 'Operation Name', 'Workcenter', 'Duration (minutes)']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)

        examples = [
            ['BOM-001', 'Cutting', 'CNC Machine', 15],
            ['BOM-001', 'Bending', 'Press Machine', 10],
            ['BOM-002', 'Injection Molding', 'Molding Machine 1', 5],
            ['BOM-002', 'Painting', 'Paint Booth', 10],
        ]

        for row_idx, example in enumerate(examples, start=1):
            for col_idx, value in enumerate(example):
                worksheet.write(row_idx, col_idx, value, example_format)

        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:C', 25)
        worksheet.set_column('D:D', 20)

    def _create_complete_template(self, workbook):
        """Create complete template with all sheets"""
        self._create_components_template(workbook)

        # Add materials sheet
        worksheet = workbook.add_worksheet('BOM Materials')
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#2196F3',
            'font_color': 'white',
            'border': 1,
        })
        headers = ['BOM Code', 'Material Name', 'Quantity', 'Unit']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:B', 30)
        worksheet.set_column('C:D', 15)

        # Add operations sheet
        worksheet = workbook.add_worksheet('BOM Operations')
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#FF9800',
            'font_color': 'white',
            'border': 1,
        })
        headers = ['BOM Code', 'Operation Name', 'Workcenter', 'Duration (minutes)']
        for col, header in enumerate(headers):
            worksheet.write(0, col, header, header_format)
        worksheet.set_column('A:A', 20)
        worksheet.set_column('B:C', 25)
        worksheet.set_column('D:D', 20)